import openpyxl

file_path = 'Veichi Fault Codes.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers}")

for row in ws.iter_rows(min_row=2, max_row=5):
    print([cell.value for cell in row])
